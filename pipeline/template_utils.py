"""قراءة قالب إكسل القسم وتعبئته بالقيم المستخرجة.

يدعم نمطين داخل نفس الورقة:
1) حقول «تسمية ← قيمة»: العمود A تسمية الحقل، والعمود B قيمته.
2) كتلة جدول: صف ترويسات (عدة أعمدة متجاورة نصية) تليه صفوف بيانات.

يدعم الخلايا المدمجة (merged cells) — يكتب في الخلية الأصلية للمدى.
"""
import json
import shutil

import openpyxl


# =====================================================================
# دعم الخلايا المدمجة
# =====================================================================

def _anchor_cell(ws, row, col):
    """
    يعيد (row, col) للخلية القابلة للكتابة.
    إذا كانت الخلية جزءاً من مدى مدموج، يعيد إحداثيات الخلية الأصلية
    (أعلى-يسار المدى) لأن openpyxl لا يسمح بالكتابة على MergedCell.
    """
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= row <= rng.max_row and
                rng.min_col <= col <= rng.max_col):
            return rng.min_row, rng.min_col
    return row, col


def _safe_write(ws, row, col, value):
    """كتابة آمنة تتعامل مع الخلايا المدمجة."""
    r, c = _anchor_cell(ws, row, col)
    try:
        ws.cell(row=r, column=c, value=value)
        return True
    except AttributeError:
        # خلية للقراءة فقط لسبب آخر — نتجاهلها بدل إيقاف المعالجة
        return False


# =====================================================================
# قراءة بنية القالب
# =====================================================================

def parse_template(path):
    """يعيد بنية القالب:
    {"sheet": name,
     "kv": [{"row": r, "label": txt}, ...],
     "table": {"header_row": r, "columns": [..]} أو null}
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    kv = []
    table = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        cells = [(c.column, str(c.value).strip())
                 for c in row if c.value not in (None, "")]
        if not cells:
            continue
        r = row[0].row
        # كتلة جدول: 3+ خلايا نصية متجاورة في صف واحد = صف ترويسات
        if table is None and len(cells) >= 3:
            table = {"header_row": r, "columns": [t for _, t in cells]}
            continue
        # داخل منطقة الجدول: صفوف بيانات نموذجية — نتجاهلها
        if table and r > table["header_row"]:
            continue
        # حقل تسمية في العمود A
        first_col, first_txt = cells[0]
        if first_col == 1:
            kv.append({"row": r, "label": first_txt})

    return {"sheet": ws.title, "kv": kv, "table": table}


def labels_of(fields):
    """قوائم جاهزة للموديل: تسميات الحقول + أعمدة الجدول."""
    kv_labels = [f["label"] for f in fields.get("kv", [])]
    tbl = fields.get("table")
    return kv_labels, (tbl["columns"] if tbl else [])


# =====================================================================
# تعبئة القالب
# =====================================================================

def fill_template(template_path, fields, values, out_path):
    """ينسخ القالب ويعبّئه.
    values = {"حقول": {label: value}, "جدول": [[v1, v2, ...], ...]}
    يكتب قيمة كل حقل في العمود B من صفّه، وصفوف الجدول تحت صف الترويسات.
    يتعامل مع الخلايا المدمجة تلقائياً.
    """
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb[fields["sheet"]] if fields["sheet"] in wb.sheetnames else wb.active

    # --- الحقول ---
    filled = values.get("حقول", {}) or {}
    for f in fields.get("kv", []):
        val = filled.get(f["label"])
        if val not in (None, ""):
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            _safe_write(ws, f["row"], 2, val)

    # --- الجدول ---
    tbl = fields.get("table")
    rows = values.get("جدول", []) or []
    if tbl and rows:
        hr = tbl["header_row"]
        ncols = len(tbl["columns"])

        # مسح بيانات نموذجية قديمة تحت الترويسات
        for r in range(hr + 1, ws.max_row + 1):
            for c in range(1, ncols + 1):
                _safe_write(ws, r, c, None)

        # كتابة الصفوف الجديدة
        for i, rowvals in enumerate(rows, start=1):
            for c, v in enumerate(rowvals[:ncols], start=1):
                _safe_write(ws, hr + i, c, v)

    wb.save(out_path)
    return out_path